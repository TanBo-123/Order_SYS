#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import openpyxl
#import matplotlib.pyplot as plt
import numpy as np


#用于matplotlib显示中文
#plt.rcParams['font.sans-serif'] = ['SimHei']

#Contents_line = 1
RANDOM_NUM = 80
PER_CAPITA_MAX = 17
PER_CAPUTA_MIN = 13

def Export_menu(menu,score,num):
    menu_list = []  #[菜品，评分，概率，选分]
    for menu_cell,score_cell in zip(menu[1:],score[1:]):
        if menu_cell.value != None:
            menu_list.append([menu_cell.value,score_cell.value,0,0])
        else:break
    menu_len = len(menu_list) - 1
    random_list = np.random.randint(0, menu_len, menu_len * RANDOM_NUM)
    print(menu_len,menu_list)
    for i in range(len(random_list)):
        menu_list[random_list[i]][2] += 1
    for i in range(menu_len):
        menu_list[i][3] = menu_list[i][1]*menu_list[i][2]/100
    #print(menu_list)
    menu_temp = menu_list.copy()
    #print(menu_temp)
    menu_list.sort(key=lambda x:x[3],reverse=True)
    print(menu_list[0:num])
    #print(menu_temp,'\n'+'\n',menu_list)
    return [menu_temp.index(menu_list[i]) for i in range(num)]

def Order_menu(people_num):
    
    pass

fpath =  './menu.xlsx'
wb = openpyxl.load_workbook(fpath)
sheets = wb.sheetnames
ws = wb[wb.sheetnames[0]]
print(len(ws["A"]), type(ws["A"]))
#for cell in ws["A"]:
#    print(type(cell),cell.value)

L = Export_menu(ws["A"],ws["C"],3)
print("L = ",L)
wb.save(fpath)